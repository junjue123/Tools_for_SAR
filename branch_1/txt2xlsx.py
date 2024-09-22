import os
import math
from openpyxl import Workbook
import time

def calculate_corners(x, y, w, h, angle):
    radians = angle  # 这里假设传入的angle已经是弧度值

    half_w = w / 2.0
    half_h = h / 2.0

    corners = [
        (-half_w, -half_h),
        (half_w, -half_h),
        (half_w, half_h),
        (-half_w, half_h)
    ]

    rotated_corners = []
    for corner in corners:
        x_rotated = x + corner[0] * math.cos(radians) - corner[1] * math.sin(radians)
        y_rotated = y + corner[0] * math.sin(radians) + corner[1] * math.cos(radians)
        rotated_corners.append([round(x_rotated, 2), round(y_rotated, 2)])

    return rotated_corners

def calculate_angle(corners):
    dx = corners[1][0] - corners[0][0]
    dy = corners[1][1] - corners[0][1]
    angle = math.atan2(dy, dx)
    return round(angle, 2)

def txt_to_xlsx(txt_file_path):
    base_path, file_name = os.path.split(txt_file_path)
    file_base_name = os.path.splitext(file_name)[0]

    xlsx_file_path = os.path.join(base_path, f"{file_base_name}.xlsx")

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='class')
    ws.cell(row=1, column=2, value='confidence')
    ws.cell(row=1, column=3, value='box')
    ws.cell(row=1, column=4, value='angle')  # 新增角度列

    with open(txt_file_path, 'r') as file:
        lines = file.readlines()

    for index, line in enumerate(lines, start=2):
        parts = line.strip().split()
        first = parts[0]
        values = [first] + list(map(float, parts[1:]))

        if len(values) == 7:
            class_value = str(values[0])
            x, y, w, h, angle = values[1:6]
            corners = calculate_corners(x, y, w, h, angle)
            box_value = f"{corners}"
            confidence_value = values[6]
            angle_value = angle  # 使用传入的角度

        elif len(values) == 10:
            class_value = str(values[0])
            corners = [[values[1], values[2]], [values[3], values[4]],
                       [values[5], values[6]], [values[7], values[8]]]
            box_value = f"{corners}"
            confidence_value = values[9]
            angle_value = calculate_angle(corners)  # 反向计算角度

        else:
            raise ValueError("TXT文件行的格式不正确。")

        ws.cell(row=index, column=1, value=class_value)
        ws.cell(row=index, column=2, value=confidence_value)
        ws.cell(row=index, column=3, value=box_value)
        ws.cell(row=index, column=4, value=angle_value)  # 写入角度值

    wb.save(xlsx_file_path)
    print(f"Excel文件已保存至: {xlsx_file_path}")
    time.sleep(1)

if __name__ == '__main__':
    # 调用函数转换JSON文件为Excel文件
    txt_to_xlsx('D:\Lic\python_zxz\licpan0821\pyqt\output\optical_Ship_Detection\Image1/output.txt')