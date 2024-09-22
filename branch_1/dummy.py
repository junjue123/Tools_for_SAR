from openpyxl import load_workbook, Workbook
from calculate import match_targets_by_distance,calculate_dynamic_weights
import os

def ReMix(remix_SAR,remix_optical):
    # 读取两个xlsx文件
    wb1 = load_workbook(filename=remix_SAR)
    wb2 = load_workbook(filename=remix_optical)

    # 获取第一个工作表
    ws1 = wb1.active
    ws2 = wb2.active

    # 初始化新的Workbook用于保存结果
    result_wb = Workbook()
    result_ws = result_wb.active

    # 设置标题行
    result_ws.cell(row=1, column=1, value='class_SAR')
    result_ws.cell(row=1, column=2, value='class_optical')
    result_ws.cell(row=1, column=3, value='confidence')

    # 根据中心坐标匹配目标
    matched_pairs = match_targets_by_distance(ws1, ws2)
    print(matched_pairs)

    # 遍历每一个匹配对，并将结果写入新的工作表
    for index, (i, j) in enumerate(matched_pairs):
        # 初始化值为 None，以便在目标没有匹配到时使用
        col1_file1, col2_file1 = None, None
        col1_file2, col2_file2 = None, None

        if i is not None:
            # 获取文件1的第一列和第三列数据
            col1_file1 = ws1.cell(row=i, column=1).value
            col2_file1 = ws1.cell(row=i, column=2).value

        if j is not None:
            # 获取文件2的第一列和第三列数据
            col1_file2 = ws2.cell(row=j, column=1).value
            col2_file2 = ws2.cell(row=j, column=2).value

        # 确保数值不为空且为数字类型，如果为空则设置为默认值0.0
        col2_file1 = float(col2_file1) if col2_file1 is not None else 0.0
        col2_file2 = float(col2_file2) if col2_file2 is not None else 0.0

        # 动态计算每一行的权重
        weight1, weight2 = calculate_dynamic_weights(col2_file1, col2_file2)

        # 计算合并后的第三列数据
        combined_col = (col2_file1 * weight1) + (col2_file2 * weight2)

        # 将结果写入新的Excel文件中，从第二行开始
        result_ws.cell(row=index + 2, column=1, value=col1_file1)
        result_ws.cell(row=index + 2, column=2, value=col1_file2)
        result_ws.cell(row=index + 2, column=3, value=combined_col)

    # 在当前路径下创建output/remix文件夹
    output_dir = os.path.join(os.getcwd(), 'output', 'remix')
    os.makedirs(output_dir, exist_ok=True)

    # 保存结果到新的xlsx文件中
    output_file = os.path.join(output_dir, 'remix_result.xlsx')
    result_wb.save(output_file)


if __name__ == '__main__':
    remix_SAR = "E:\SAR_projects\pyqt_820/test\SAR.xlsx"
    remix_optical = "E:\SAR_projects\pyqt_820/test\optical.xlsx"
    ReMix(remix_SAR,remix_optical)